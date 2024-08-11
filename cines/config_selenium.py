import os
import re
import time
import requests
import openpyxl
import datetime
from datetime import date
from openpyxl import Workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains


def save_excel(cinema,current_date):
    if not os.path.exists("results"):
        os.makedirs("results")
    
    current_hour_ = datetime.datetime.now().strftime('%H-%M-%S')

    distin_name = "results/"+cinema+"-"+str(current_date) + \
        " "+str(current_hour_)+".xlsx"
    
    workbook.save(filename=distin_name)
    print("Información de los cines guardada en el archivo Excel.")
    
def close_popup(selector_type,element):
    try:
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((selector_type, element)))
        # Cerrar el cuadro de diálogo emergente
        close_button = driver.find_element((selector_type, element))
        close_button.click()
    except Exception as e:
        print(f"No hay pop-up para cerrar")

def write_movie_data(sheet, date, country, cinema_brand, cinema_location, movie_title, time, classification, language):
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1).value = date
    sheet.cell(row=next_row, column=2).value = country
    sheet.cell(row=next_row, column=3).value = cinema_brand
    sheet.cell(row=next_row, column=4).value = cinema_location
    sheet.cell(row=next_row, column=5).value = movie_title
    sheet.cell(row=next_row, column=6).value = time
    sheet.cell(row=next_row, column=7).value = classification
    sheet.cell(row=next_row, column=8).value = language

# Ruta al ejecutable de EdgeDriver ( especificar la ruta correcta )

paths = [
    'C:/Users/ve180/Downloads/edgedriver_win64/msedgedriver.exe',
    'C:/Users/Wizar/Documents/Curso de Business Intelligence y Big Data Febrero 2024/Proyecto Final/msedgedriver.exe',
    ]

# Configurar las opciones de Edge
options = Options()
options.add_argument("--inprivate")
options.add_argument("--start-maximized")
# Esta opción mantiene la pestaña abierta después de que el script termine
options.detach = True

# Configurar el servicio de EdgeDriver
service = Service(executable_path=paths[1])

# Crear el driver de Edge con las opciones y el servicio configurados
driver = webdriver.Edge(service=service, options=options)

# Crear libro de trabajo, hoja y fila de encabezados
workbook = openpyxl.Workbook()
sheet = workbook.active

next_row = 1

sheet.cell(row=next_row, column=1).value = "Fecha"
sheet.cell(row=next_row, column=2).value = "Pais"
sheet.cell(row=next_row, column=3).value = "Cine"
sheet.cell(row=next_row, column=4).value = "Nombre Cine"
sheet.cell(row=next_row, column=5).value = "Titulo"
sheet.cell(row=next_row, column=6).value = "Hora"
sheet.cell(row=next_row, column=7).value = "Formato"
sheet.cell(row=next_row, column=8).value = "Idioma"