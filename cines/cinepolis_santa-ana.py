import openpyxl
import datetime
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from webdriver_manager.microsoft import EdgeChromiumDriverManager

def close_popup():
    try:
        WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.ID, 'takeover-close')))

        # Cerrar el cuadro de diálogo emergente
        close_button = driver.find_element(By.ID, 'takeover-close')
        close_button.click()
    except Exception as e:
        print(f"No hay pop-up para cerrar")

def write_movie_data(sheet, date, cinema_brand, cinema_location, movie_title, schedules, language, classification):
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = date
        sheet.cell(row=next_row, column=2).value = cinema_brand
        sheet.cell(row=next_row, column=3).value = cinema_location
        sheet.cell(row=next_row, column=4).value = movie_title
        sheet.cell(row=next_row, column=5).value = "".join(schedules)
        sheet.cell(row=next_row, column=6).value = language
        sheet.cell(row=next_row, column=7).value = classification

def scrape_movie_details(movie_listing):
    # Extraer el título de la película
    movie_title_element = movie_listing.find_element(By.CSS_SELECTOR, ".informacion-general h3")
    movie_title = movie_title_element.text.strip()

    # Extraer los horarios (si es necesario, ajustar el selector)
    schedule_elements = movie_listing.find_elements(By.CSS_SELECTOR, ".horas a p")
    schedules = [schedule.text.strip() for schedule in schedule_elements]

    # Extraer el idioma (si es necesario, ajustar el selector)
    language_element = movie_listing.find_element(By.CSS_SELECTOR, ".formato p")
    language = language_element.text.strip()

    # Extraer la clasificación (asumiendo que está en la clase "extra-hover")
    classification_element = movie_listing.find_element(By.CLASS_NAME, "extra-hover")
    classification = classification_element.text.strip()

    # Extraer los formatos y sus horarios
    formats_with_times = {}
    format_elements = movie_listing.find_elements(By.CSS_SELECTOR, ".formato")
    for format_element in format_elements:
        format_name = format_element.find_element(By.CSS_SELECTOR, ".formato-nombre").text.strip()
        times = [a.text.strip() for a in format_element.find_elements(By.CSS_SELECTOR, ".horas a p")]
        formats_with_times[format_name] = times

    return movie_title, schedules, classification, formats_with_times

# Creando libro de trabajo, hoja y fila de encabezados
workbook = openpyxl.Workbook()
sheet = workbook.active

next_row = 1

sheet.cell(row=next_row, column=1).value = "Fecha"
sheet.cell(row=next_row, column=2).value = "Cine"
sheet.cell(row=next_row, column=3).value = "Nombre Cine"
sheet.cell(row=next_row, column=4).value = "Titulo"
sheet.cell(row=next_row, column=5).value = "Hora"
sheet.cell(row=next_row, column=6).value = "Idioma"
sheet.cell(row=next_row, column=7).value = "Formato"

# Crear una nueva instancia de Options
edge_options = Options()
edge_options.add_argument("--inprivate")
edge_options.add_argument("--start-maximized")

# Configurar el servicio Edge y crear el controlador
service = Service(EdgeChromiumDriverManager().install())
driver = webdriver.Edge(service=service, options=edge_options)
driver.implicitly_wait(5)

# Navegar a la página web
driver.get('https://cinepolis.com.sv/')

close_popup()

ciudad_element = driver.find_element
ciudad_options = driver.find_element(By.ID,'ciudad')
select=Select(ciudad_options)

try:
    select.select_by_index(2)
except NoSuchElementException:
    print('\nError!!!!!!!!!\nThe item does not exist')

button_element = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
button_element.click()

cinema_element = driver.find_element(By.ID, "cinema-245")
cinema_name = cinema_element.find_element(By.TAG_NAME, "h2").text
cinema_name_parts = cinema_name.split(" ")
cinema_brand = cinema_name_parts[0].capitalize()
cinema_location = " ".join(cinema_name_parts[1:]).title()

# Buscar el elemento que contiene las películas
list_movies = driver.find_element(By.CLASS_NAME, "movies")

# Iterar sobre cada listado de películas
for movie_listing in list_movies.find_elements(By.CSS_SELECTOR, ".SingleScheduleMovie__SingleScheduleComponent-sc-1n3hti2-0"):
    
    # Extraer los detalles de la película
    movie_title, schedules, classification, formats_with_times = scrape_movie_details(movie_listing)

    # Escribir los datos de la película en la hoja de Excel para cada horario
    for format_name, times in formats_with_times.items():
        for time in times:
            # Escribir los datos en el Excel:
            write_movie_data(sheet, datetime.date.today().strftime('%d-%m-%Y'), cinema_brand, cinema_location, movie_title, time, format_name,classification)

# Guardar el libro de trabajo de Excel
workbook.save(filename="cinemas.xlsx")

# Autoajustar el ancho de todas las columnas
for column_cells in sheet.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    sheet.column_dimensions[column_cells[1].column_letter].width = length

print("Información de los cines guardada en el archivo Excel.")

# Salir del controlador
driver.quit()
