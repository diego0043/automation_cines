import datetime
import os
import pandas as pd
from openpyxl import load_workbook

def find_excel_files(directory, keyword):
    try:
        """Encuentra todos los archivos Excel en el directorio especificado que coincidan con la palabra clave."""
        return [os.path.join(directory, f) for f in os.listdir(directory) if f.endswith('.xlsx') and keyword in f]
    except Exception as e:
        print(f"Error al buscar archivos Excel: {e}")

def read_excel_file(file):
    """Lee un archivo Excel y devuelve un DataFrame con las primeras 8 columnas."""
    wb = load_workbook(filename=file, data_only=True)
    ws = wb.active

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # Salta la primera fila (encabezado)
        data.append(row[:8])  # Solo toma las primeras 8 columnas

    df = pd.DataFrame(data)
    return df

def unify_excels(directory, keyword, output_file, header):
    """Unifica todos los archivos Excel en el directorio especificado que coincidan con la palabra clave."""
    try:
        files = find_excel_files(directory, keyword)

        dfs = []
        for file in files:
            df = read_excel_file(file)
            if len(df.columns) != len(header):
                print(f"El archivo {file} no tiene exactamente 8 columnas. Se omite este archivo.")
                continue
            dfs.append(df)

        if not dfs:
            print("No se encontraron archivos con el número correcto de columnas.")
            return

        combined_df = pd.concat(dfs, ignore_index=True)
        combined_df.columns = header
        combined_df.to_excel(os.path.join(directory, output_file), index=False)

        print(f"Archivos unificados exitosamente en '{output_file}'")
    except Exception as e:
        print(f"Error al unificar archivos Excel: {e}")

# Directorio donde se encuentran los archivos Excel
directory = 'results'

# Palabra clave para buscar en los nombres de los archivos Excel
keyword = 'cinepolis'

# Archivo de salida
output_file = 'archivo_unificado_'+keyword+'_'+datetime.date.today().strftime('%d-%m-%Y')+'.xlsx'

# Encabezado correcto
header = ['Fecha', 'Pais', 'Cine', 'Nombre Cine', 'Titulo', 'Hora', 'Formato', 'Idioma']

# Unificar los archivos
unify_excels(directory, keyword, output_file, header)
